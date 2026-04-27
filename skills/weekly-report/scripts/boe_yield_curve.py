#!/usr/bin/env python3
"""Bank of England UK gilt nominal spot yield curve (sync port).

Downloads BOE current-month and historical XLSX zips, parses the "4. spot curve"
sheet, merges current overrides over historical, filters by [start, end] window
and tenors, and emits a JSON document on stdout.
"""

from __future__ import annotations

import argparse
import io
import math
import re
import sys
import zipfile
from datetime import date
from typing import Any

from _common import add_window_args, build_window_payload, emit, fail, parse_date, resolve_window

CURRENT_ZIP_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/"
    "latest-yield-curve-data.zip"
)
HISTORICAL_ZIP_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/"
    "glcnominalddata.zip"
)
CURRENT_XLSX_NAME = "GLC Nominal daily data current month.xlsx"
HISTORICAL_XLSX_PATTERN = re.compile(
    r"GLC Nominal daily data_\d{4} to present\.xlsx$",
    flags=re.IGNORECASE,
)
SHEET_NAME = "4. spot curve"
DOWNLOAD_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/zip,application/octet-stream,*/*",
    "Accept-Language": "en-GB,en;q=0.9",
    "Referer": "https://www.bankofengland.co.uk/",
}


def safe_float(v: Any) -> float | None:
    try:
        n = float(v)
        if math.isnan(n) or math.isinf(n):
            return None
        return n
    except Exception:
        return None


def round2(v: float) -> float:
    r = round(float(v), 2)
    return 0.0 if r == -0.0 else r


def to_iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "date"):
        try:
            return value.date().isoformat()
        except Exception:
            return None
    if isinstance(value, date):
        return value.isoformat()
    try:
        return date.fromisoformat(str(value)[:10]).isoformat()
    except Exception:
        return None


def to_tenor_label(years: float) -> str:
    if years <= 0:
        return "0Y"
    if years < 1:
        months = max(1, int(round(years * 12)))
        return f"{months}M"
    if abs(years - round(years)) < 1e-8:
        return f"{int(round(years))}Y"
    return f"{years:.1f}Y"


def normalize_tenor_filter(tenors: list[str]) -> set[str]:
    out: set[str] = set()
    for raw in tenors:
        text = str(raw or "").strip().upper()
        if not text:
            continue
        if text.endswith("Y") or text.endswith("M"):
            out.add(text)
            continue
        try:
            num = float(text)
        except Exception:
            continue
        if num < 1:
            out.add(f"{max(1, int(round(num * 12)))}M")
        elif abs(num - round(num)) < 1e-8:
            out.add(f"{int(round(num))}Y")
        else:
            out.add(f"{num:.1f}Y")
    return out


def tenor_sort_key(tenor: str) -> tuple[int, float]:
    text = str(tenor or "").strip().upper()
    if text.endswith("M"):
        try:
            return (0, float(text[:-1]))
        except Exception:
            return (0, 9999.0)
    if text.endswith("Y"):
        try:
            return (1, float(text[:-1]))
        except Exception:
            return (1, 9999.0)
    return (2, 9999.0)


def download_xlsx_from_zip(
    zip_url: str,
    exact_name: str | None = None,
    name_pattern: re.Pattern[str] | None = None,
) -> tuple[str, bytes]:
    try:
        import requests
    except Exception as exc:
        raise RuntimeError("requests is required") from exc

    response = requests.get(zip_url, headers=DOWNLOAD_HEADERS, timeout=30, allow_redirects=True)
    if response.status_code == 403:
        response = requests.get(
            zip_url,
            headers={**DOWNLOAD_HEADERS, "Cache-Control": "no-cache", "Pragma": "no-cache"},
            timeout=30,
            allow_redirects=True,
        )
    if response.status_code >= 400:
        raise RuntimeError(
            f"download failed status={response.status_code} url={zip_url}"
        )

    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        for name in archive.namelist():
            base = name.split("/")[-1]
            if exact_name and base == exact_name:
                return base, archive.read(name)
            if name_pattern and name_pattern.search(base):
                return base, archive.read(name)
    expected = exact_name or (name_pattern.pattern if name_pattern else ".xlsx")
    raise FileNotFoundError(f"{expected} not found in BOE zip payload: {zip_url}")


def parse_spot_curve_rows(xlsx_bytes: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is required to parse BOE xlsx") from exc

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Worksheet '{SHEET_NAME}' not found")
    sheet = workbook[SHEET_NAME]

    tenor_cells = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    tenors: list[tuple[int, str]] = []
    for idx, value in enumerate(tenor_cells):
        if idx == 0:
            continue
        number = safe_float(value)
        if number is None:
            continue
        tenors.append((idx, to_tenor_label(number)))
    if not tenors:
        raise RuntimeError("No tenor headers found in BOE worksheet")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=6, values_only=True):
        row_date = values[0] if values else None
        iso = to_iso_date(row_date)
        if not iso:
            continue
        for col_idx, tenor in tenors:
            if col_idx >= len(values):
                continue
            v = safe_float(values[col_idx])
            if v is None:
                continue
            rows.append({"date": iso, "tenor": tenor, "value": round2(v)})
    return rows


def coverage(rows: list[dict[str, Any]]) -> tuple[date | None, date | None]:
    dates: list[date] = []
    for r in rows:
        try:
            dates.append(date.fromisoformat(str(r.get("date") or "")))
        except Exception:
            continue
    if not dates:
        return None, None
    return min(dates), max(dates)


def merge_rows(base_rows, override_rows):
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for r in base_rows:
        key = (str(r.get("date") or ""), str(r.get("tenor") or ""))
        if key[0] and key[1]:
            merged[key] = r
    for r in override_rows:
        key = (str(r.get("date") or ""), str(r.get("tenor") or ""))
        if key[0] and key[1]:
            merged[key] = r
    return list(merged.values())


def main() -> None:
    parser = argparse.ArgumentParser(description="BOE UK gilt nominal spot yield curve.")
    add_window_args(parser, default_days=7)
    parser.add_argument("--tenors", default="2Y,5Y,10Y,30Y", help="Comma-separated tenors filter")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    start, end = resolve_window(args.start_date, args.end_date, days=args.days)
    start_dt = date.fromisoformat(start)
    tenor_list = [t.strip() for t in args.tenors.split(",") if t.strip()]
    tenor_filter = normalize_tenor_filter(tenor_list)
    if tenor_list and not tenor_filter:
        fail("Invalid tenors format", tenors=args.tenors)

    trace: list[dict[str, Any]] = []
    current_rows: list[dict[str, Any]] = []
    current_error: str | None = None
    current_file: str | None = None
    try:
        current_file, xlsx_bytes = download_xlsx_from_zip(CURRENT_ZIP_URL, exact_name=CURRENT_XLSX_NAME)
        current_rows = parse_spot_curve_rows(xlsx_bytes)
        trace.append({"stage": "current_ok", "rows": len(current_rows)})
    except Exception as exc:
        current_error = str(exc)
        trace.append({"stage": "current_failed", "error": current_error})

    cmin, cmax = coverage(current_rows)
    need_historical = bool(current_error) or not current_rows or (cmin is not None and start_dt < cmin)

    historical_rows: list[dict[str, Any]] = []
    historical_error: str | None = None
    historical_file: str | None = None
    if need_historical:
        try:
            historical_file, hist_bytes = download_xlsx_from_zip(
                HISTORICAL_ZIP_URL, name_pattern=HISTORICAL_XLSX_PATTERN
            )
            historical_rows = parse_spot_curve_rows(hist_bytes)
            trace.append({"stage": "historical_ok", "rows": len(historical_rows)})
        except Exception as exc:
            historical_error = str(exc)
            trace.append({"stage": "historical_failed", "error": historical_error})

    merged = merge_rows(historical_rows, current_rows)
    if not merged:
        fail(
            "Failed to fetch BOE yield curve data",
            current_source_error=current_error,
            historical_source_error=historical_error,
            debug_trace=trace if args.debug else None,
        )

    filtered: list[dict[str, Any]] = []
    for row in merged:
        d = str(row.get("date") or "")
        if d < start or d > end:
            continue
        tenor = str(row.get("tenor") or "")
        if tenor_filter and tenor not in tenor_filter:
            continue
        filtered.append(row)
    filtered.sort(key=lambda x: (str(x["date"]), tenor_sort_key(str(x["tenor"]))))

    series_map: dict[str, list[dict[str, Any]]] = {}
    for row in filtered:
        series_map.setdefault(str(row["tenor"]), []).append({"date": row["date"], "value": row["value"]})

    sources_used: list[dict[str, Any]] = []
    if current_rows:
        cmin, cmax = coverage(current_rows)
        sources_used.append(
            {
                "source": "boe_current_month",
                "url": CURRENT_ZIP_URL,
                "file": current_file or CURRENT_XLSX_NAME,
                "coverage": {
                    "start_date": cmin.isoformat() if cmin else None,
                    "end_date": cmax.isoformat() if cmax else None,
                },
                "rows": len(current_rows),
            }
        )
    if historical_rows:
        hmin, hmax = coverage(historical_rows)
        sources_used.append(
            {
                "source": "boe_historical",
                "url": HISTORICAL_ZIP_URL,
                "file": historical_file or "",
                "coverage": {
                    "start_date": hmin.isoformat() if hmin else None,
                    "end_date": hmax.isoformat() if hmax else None,
                },
                "rows": len(historical_rows),
            }
        )

    data_gaps: list[str] = []
    if current_error:
        data_gaps.append(f"current-month source failed: {current_error}")
    if need_historical and historical_error:
        data_gaps.append(f"historical source failed: {historical_error}")

    payload: dict[str, Any] = {
        "section": "fixed_income_uk",
        "source": "Bank of England",
        "source_url": CURRENT_ZIP_URL,
        "sheet": SHEET_NAME,
        "window": build_window_payload(start, end),
        "unit": "percent",
        "sources_used": sources_used,
        "fallback_applied": need_historical,
        "tenors": sorted(series_map.keys(), key=tenor_sort_key),
        "row_count": len(filtered),
        "rows": filtered,
        "series": [
            {"tenor": tenor, "observations": obs}
            for tenor, obs in sorted(series_map.items(), key=lambda x: tenor_sort_key(x[0]))
        ],
        "data_gaps": data_gaps,
    }
    if args.debug:
        payload["debug_trace"] = trace
    emit(payload)


if __name__ == "__main__":
    main()
